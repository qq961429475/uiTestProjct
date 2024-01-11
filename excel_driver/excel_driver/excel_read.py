'''
    Excel数据驱动效果实现。
        实现目的：基于excel中的内容，来调用关键字类实现自动化测试的执行。
        相当于excel文件就是一个测试用例，底层代码就是关键字驱动类以及excel驱动类
'''
import openpyxl
from selenium import webdriver

# 获取到excel，进入sheet页中，读取单元格内容
from common.utils_copy import BasePage

driver1 = webdriver.Chrome()
driver = BasePage(driver1)

wb = openpyxl.load_workbook('../excel_data/data.xlsx')
# 获取所有的sheet页
sheets = wb.sheetnames
# 遍历所有sheet页
for sheet in sheets:
    # items为一行数据
    items = wb[sheet]
    # 如果sheet中包含特定字段，直接continue
    # 遍历sheet页中所有的单元格
    for values in items.values:
        # 1. 读取用例的执行部分的内容。
        if type(values[0]) is int:
            '''
                用例结构：
                    1. 编号：不要管
                    2. 调用的关键字，结合关键字驱动类应用
                    3. 关键字对应的参数：测试数据，结合关键字驱动类应用
                    4. 本条用例的行为描述：行为记录，可以添加到日志中进行输出
            '''
            # 提取本行的测试数据
            data = {}
            data['name'] = values[2]
            data['value'] = values[3]
            data['txt'] = values[4]
            # 优化测试数据内容,将所有为None的数据全部清除出data中
            for key in list(data.keys()):
                # print(key)
                if data[key] is None:
                    del data[key]
            print(data)

            # 断言可能不会只有一种，只要有assert关键字，就是一个断言函数
            if 'assert' in values[1]:
                # 只有断言函数才会有返回值。
                status = getattr(driver, values[1])(**data)
                # 基于status来写入测试的结果
                if status is True:
                    # 执行Pass写入
                    items.cell(row=values[0] + 2, column=7).value = 'Pass'
                else:
                    # 执行Failed写入
                    items.cell(row=values[0] + 2, column=7).value = 'Failed'
            else:
                getattr(driver, values[1])(**data)

# 执行excel的保存
wb.save('../excel_data/data.xlsx')
