import openpyxl

# excel数据读取：如果在操作时方法无法点出来，就手动填写。
# 1. 读取excel文件。
excel = openpyxl.load_workbook('../excel_data/data.xlsx')
# 2. 读取指定sheet页
sheet = excel['Sheet1']
# 3. 读取指定的单元格
# cell = sheet['A1']
# print(cell.value)
# 读取整个sheet页内容：没有内容的单元格通过None值来表示。
# valus = sheet.values
# for v in valus:
#     print(v)  # v是元组的格式，通过下标可以获取到指定单元格的内容。
# 获取所有的sheet页中的单元格内容。
# 1. 获取excel中所有的sheet名称，返回list格式的内容。
# sheets = excel.sheetnames
# print(sheets)
# # 2. 通过循环读取sheets中内容，赋值给excel以key的形式
# for sheet in sheets:
#     print('*' * 20)
#     # 3. 每一个sheet中的单元格内容输出
#     for value in excel[sheet].values:
#         print(value)
# excel的写入:要先关闭excel文件，记得写完后执行保存的操作。
sheet.cell(row=5, column=5).value = '虚竹啥也不缺，就缺钱'
excel.save('../excel_data/data.xlsx')
