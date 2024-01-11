from datetime import datetime

from openpyxl import load_workbook
from selenium import webdriver

from utils_copy import BasePage


def excel_test(filename):
    global driver
    global title
    wb = load_workbook(filename)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        for items in ws.values:
            print(items)
            if type(items[0]) is int:
                data = [items[2], items[3], items[4]]
                data = [i for i in data if i is not None]
                print(data)
                if items[1] == 'visit':
                    driver = BasePage(webdriver.Chrome())
                    driver.visit(data[0])
                elif items[1] == '用例标题':
                    title = items[2]
                    continue
                elif 'assert' in items[1]:
                    status = getattr(driver, items[1])(*data)
                    if status is True:
                        ws.cell(row=items[0] + 2, column=7).value = 'pass'
                    else:
                        now = datetime.now()
                        formatted_date = now.strftime("%Y-%m-%d_%H_%M_%S")
                        image_name = title + '_' + formatted_date + '.png'
                        driver.screen_shot(f'../images/{image_name}')
                        ws.cell(row=items[0] + 2, column=7).value = 'fail'
                else:
                    getattr(driver, items[1])(*data)
        wb.save('../data/data_done.xlsx')


if __name__ == '__main__':
    excel_test('../data/data.xlsx')
